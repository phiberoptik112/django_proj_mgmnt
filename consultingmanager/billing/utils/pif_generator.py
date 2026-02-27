"""
PIF Generator Utility
Extracts data from proposal documents (.doc/.docx/.txt) and generates PIF Excel files.
"""

import logging
import re
from typing import Dict, Any, List, Tuple, Optional
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import datetime
import os

# Import existing utilities
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../..'))

from files.utils.quick_project_analysis import extract_text_from_file
from projects.utils.proposal_parser import ProposalParser

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available. Install it with: pip install openpyxl")

logger = logging.getLogger(__name__)


class PIFFieldExtractor:
    """
    Extracts PIF fields from proposal text using regex patterns and context analysis.
    """

    # Regex patterns for field extraction
    PATTERNS = {
        'project_number': [
            (r'\b(P?\d{2}-\d{3})\b', 1.0),  # Exact pattern P23-001 or 23-001
            (r'Project\s+(?:No|Number|#)[:\s]+([A-Z]?\d{2}-\d{3})', 0.9),
        ],
        'project_name': [
            (r'RE:\s+(.+?)(?:\n|$)', 0.95),  # After RE: line
            (r'Project:\s+(.+?)(?:\n|$)', 0.8),
        ],
        'client_name': [
            (r'(?:Client|Company)[:\s]+(.+?)(?:\n|$)', 0.9),
        ],
        'billing_contact': [
            (r'^([A-Z][a-z]+\s+[A-Z][a-z]+)\s*$', 0.7),  # Name pattern (first occurrence)
        ],
        'billing_contact_email': [
            (r'\b([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b', 0.95),
        ],
        'phone': [
            (r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', 0.9),
        ],
        'fee_contract_amount': [
            (r'lump\s+sum\s+fee\s+of\s+\$?([\d,]+(?:\.\d{2})?)', 1.0),
            (r'total\s+fee[:\s]+\$?([\d,]+(?:\.\d{2})?)', 0.9),
            (r'contract\s+amount[:\s]+\$?([\d,]+(?:\.\d{2})?)', 0.9),
            (r'\$\s*([\d,]+(?:\.\d{2})?)', 0.5),  # Any dollar amount (low confidence)
        ],
        'expenses': [
            (r'expenses?[:\s]+\$?([\d,]+(?:\.\d{2})?)', 0.9),
            (r'reimbursable[:\s]+\$?([\d,]+(?:\.\d{2})?)', 0.8),
        ],
        'project_start_date': [
            (r'(?:start|commence)\s+(?:date)?[:\s]+((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4})', 0.95),
            (r'(?:start|commence)[:\s]+(\d{1,2}/\d{1,2}/\d{4})', 0.9),
        ],
        'date_entered': [
            (r'^((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4})', 0.95),  # First date in doc
            (r'^(\d{1,2}/\d{1,2}/\d{4})', 0.9),
        ],
        'project_manager': [
            (r'Sincerely,?\s*\n+([A-Z][a-z]+\s+[A-Z][a-z]+)', 0.85),  # After signature
            (r'Project\s+Manager[:\s]+(.+?)(?:\n|$)', 0.95),
        ],
        'originator': [
            (r'Prepared\s+by[:\s]+(.+?)(?:\n|$)', 0.9),
            (r'Author[:\s]+(.+?)(?:\n|$)', 0.8),
        ],
        'project_location_city': [
            (r'Location[:\s]+([A-Za-z\s]+),\s+[A-Z]{2}', 0.9),  # City, ST format
        ],
        'project_location_state': [
            (r'Location[:\s]+[A-Za-z\s]+,\s+([A-Z]{2})', 0.9),  # City, ST format
        ],
        'purchase_order_number': [
            (r'P\.?O\.?\s+(?:Number|#)[:\s]+([A-Z0-9-]+)', 0.95),
            (r'Purchase\s+Order[:\s]+([A-Z0-9-]+)', 0.95),
        ],
    }

    # Phase keywords for billing phase extraction
    PHASE_KEYWORDS = {
        'programming': ['programming', 'program'],
        'schematic_design': ['schematic', 'schematic design'],
        'design_development': ['design development', 'dd phase'],
        'construction_documents': ['construction documents', 'cd phase', 'construction drawings'],
        'bidding_negotiation': ['bidding', 'negotiation', 'bid phase'],
        'construction_administration': ['construction administration', 'ca phase'],
        'professional_services': ['professional services', 'consulting services'],
    }

    def __init__(self, text: str, file_path: Optional[str] = None):
        """
        Initialize the field extractor

        Args:
            text: Extracted proposal text
            file_path: Optional path to original file for filename-based extraction
        """
        self.text = text
        self.file_path = file_path
        self.lines = text.split('\n')
        self.extracted_data = {}
        self.confidence_scores = {}

    def extract_all_fields(self) -> Tuple[Dict[str, Any], Dict[str, float]]:
        """
        Extract all PIF fields from the proposal text

        Returns:
            Tuple of (extracted_data dict, confidence_scores dict)
        """
        # Extract each field
        for field_name in self.PATTERNS.keys():
            value, confidence = self._extract_field(field_name)
            if value:
                self.extracted_data[field_name] = value
                self.confidence_scores[field_name] = confidence

        # Extract recipient block (client name and billing contact)
        self._extract_recipient_block()

        # Extract project name from RE: line if not found
        if 'project_name' not in self.extracted_data:
            self._extract_project_name_from_re_line()

        # Extract project number from filename if not found
        if 'project_number' not in self.extracted_data and self.file_path:
            self._extract_project_number_from_filename()

        # Set defaults for common fields
        self._set_defaults()

        logger.info(f"Extracted {len(self.extracted_data)} fields with average confidence {self._calculate_avg_confidence():.2f}")

        return self.extracted_data, self.confidence_scores

    def _extract_field(self, field_name: str) -> Tuple[Optional[Any], float]:
        """
        Extract a single field using regex patterns

        Args:
            field_name: Name of the field to extract

        Returns:
            Tuple of (value, confidence_score)
        """
        if field_name not in self.PATTERNS:
            return None, 0.0

        patterns = self.PATTERNS[field_name]
        best_value = None
        best_confidence = 0.0

        for pattern, base_confidence in patterns:
            matches = re.finditer(pattern, self.text, re.MULTILINE | re.IGNORECASE)
            for match in matches:
                value = match.group(1) if match.groups() else match.group(0)
                value = value.strip()

                # Calculate context-aware confidence
                confidence = self._calculate_confidence(
                    field_name, value, base_confidence, match.start()
                )

                if confidence > best_confidence:
                    best_value = value
                    best_confidence = confidence

        # Format value based on field type
        if best_value:
            best_value = self._format_value(field_name, best_value)

        return best_value, best_confidence

    def _calculate_confidence(self, field_name: str, value: str, base_confidence: float, position: int) -> float:
        """
        Calculate context-aware confidence score

        Args:
            field_name: Name of the field
            value: Extracted value
            base_confidence: Base confidence from pattern match
            position: Position in text where value was found

        Returns:
            Adjusted confidence score (0.0 to 1.0)
        """
        # Start with pattern confidence (30%)
        pattern_score = base_confidence * 0.3

        # Context relevance (30%)
        context = self.text[max(0, position-100):min(len(self.text), position+100)].lower()
        context_score = 0.0

        # Check for field-specific keywords near the value
        context_keywords = {
            'billing_contact_email': ['email', 'billing', 'contact'],
            'phone': ['phone', 'tel', 'telephone', 'contact'],
            'fee_contract_amount': ['fee', 'contract', 'amount', 'lump sum'],
            'expenses': ['expense', 'reimbursable'],
            'project_start_date': ['start', 'commence', 'begin'],
            'project_manager': ['manager', 'pm', 'sincerely'],
        }

        if field_name in context_keywords:
            keyword_matches = sum(1 for kw in context_keywords[field_name] if kw in context)
            context_score = min(1.0, keyword_matches / len(context_keywords[field_name])) * 0.3
        else:
            context_score = 0.15  # Default if no specific keywords

        # Format validity (20%)
        format_score = self._validate_format(field_name, value) * 0.2

        # Position heuristic (20%)
        position_score = self._score_position(field_name, position) * 0.2

        total_score = pattern_score + context_score + format_score + position_score
        return min(1.0, total_score)

    def _validate_format(self, field_name: str, value: str) -> float:
        """
        Validate format of extracted value

        Returns:
            Validation score (0.0 to 1.0)
        """
        try:
            if field_name in ['fee_contract_amount', 'expenses']:
                # Try to convert to Decimal
                cleaned = value.replace(',', '').replace('$', '')
                Decimal(cleaned)
                return 1.0
            elif field_name in ['project_start_date', 'date_entered']:
                # Try to parse as date
                # Accept common formats
                if re.match(r'\d{1,2}/\d{1,2}/\d{4}', value):
                    return 1.0
                if re.match(r'(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}', value, re.IGNORECASE):
                    return 1.0
                return 0.5
            elif field_name == 'billing_contact_email':
                # Validate email format
                if '@' in value and '.' in value.split('@')[1]:
                    return 1.0
                return 0.3
            elif field_name == 'phone':
                # Validate phone format
                digits = re.sub(r'\D', '', value)
                if len(digits) == 10:
                    return 1.0
                return 0.5
            else:
                # Generic validation - just check it's not empty
                return 1.0 if value else 0.0
        except Exception:
            return 0.0

    def _score_position(self, field_name: str, position: int) -> float:
        """
        Score based on expected position in document

        Args:
            field_name: Name of the field
            position: Position in text

        Returns:
            Position score (0.0 to 1.0)
        """
        text_length = len(self.text)
        relative_position = position / text_length if text_length > 0 else 0.5

        # Expected positions for certain fields
        if field_name in ['date_entered', 'billing_contact', 'client_name']:
            # Expected at top (first 20%)
            return 1.0 if relative_position < 0.2 else 0.5
        elif field_name in ['project_manager']:
            # Expected at bottom (last 30%)
            return 1.0 if relative_position > 0.7 else 0.5
        elif field_name in ['fee_contract_amount', 'expenses']:
            # Expected in middle/bottom half
            return 1.0 if relative_position > 0.3 else 0.7
        else:
            # No strong position preference
            return 0.8

    def _format_value(self, field_name: str, value: str) -> Any:
        """
        Format extracted value to appropriate type

        Args:
            field_name: Name of the field
            value: Raw extracted value

        Returns:
            Formatted value
        """
        try:
            if field_name in ['fee_contract_amount', 'expenses']:
                # Convert to Decimal
                cleaned = value.replace(',', '').replace('$', '').strip()
                return str(Decimal(cleaned))
            elif field_name in ['project_start_date', 'date_entered']:
                # Keep as string for now, will be parsed by Django form
                return value
            else:
                # Return as string
                return value.strip()
        except Exception as e:
            logger.warning(f"Error formatting {field_name} value '{value}': {e}")
            return value

    def _extract_recipient_block(self):
        """
        Extract client name and billing contact from recipient block
        Typically appears between date and RE: line
        """
        # Find date line and RE: line
        date_line_idx = None
        re_line_idx = None

        for i, line in enumerate(self.lines):
            line_stripped = line.strip()
            # Look for date pattern at beginning
            if not date_line_idx and re.match(r'^(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}', line_stripped, re.IGNORECASE):
                date_line_idx = i
            # Look for RE: line
            if line_stripped.startswith('RE:') or line_stripped.startswith('Re:'):
                re_line_idx = i
                break

        # Extract lines between date and RE:
        if date_line_idx is not None and re_line_idx is not None and re_line_idx > date_line_idx:
            recipient_lines = []
            for i in range(date_line_idx + 1, re_line_idx):
                line = self.lines[i].strip()
                if line and len(line) > 2:  # Skip empty lines and very short lines
                    recipient_lines.append(line)

            # First non-empty line is typically billing contact (person name)
            # Second line is typically client name (company)
            if len(recipient_lines) >= 1:
                billing_contact = recipient_lines[0]
                # Check if it looks like a person name
                if re.match(r'^[A-Z][a-z]+\s+[A-Z][a-z]+', billing_contact):
                    self.extracted_data['billing_contact'] = billing_contact
                    self.confidence_scores['billing_contact'] = 0.85

            if len(recipient_lines) >= 2:
                client_name = recipient_lines[1]
                self.extracted_data['client_name'] = client_name
                self.confidence_scores['client_name'] = 0.8

    def _extract_project_name_from_re_line(self):
        """Extract project name from RE: line"""
        for line in self.lines:
            if line.strip().startswith('RE:') or line.strip().startswith('Re:'):
                # Extract everything after RE:
                project_name = re.sub(r'^RE:\s*', '', line.strip(), flags=re.IGNORECASE)
                if project_name:
                    self.extracted_data['project_name'] = project_name
                    self.confidence_scores['project_name'] = 0.9
                break

    def _extract_project_number_from_filename(self):
        """Extract project number from filename"""
        if not self.file_path:
            return

        filename = Path(self.file_path).stem
        match = re.search(r'\b(P?\d{2}-\d{3})\b', filename)
        if match:
            self.extracted_data['project_number'] = match.group(1)
            self.confidence_scores['project_number'] = 0.75  # Lower confidence from filename

    def _set_defaults(self):
        """Set default values for common fields if not extracted"""
        defaults = {
            'dlaa_office': ('Kailua', 0.5),
            'type_of_contract': ('Lump Sum', 0.5),
            'retainer_received': ('no', 0.5),
        }

        for field, (default_value, confidence) in defaults.items():
            if field not in self.extracted_data:
                self.extracted_data[field] = default_value
                self.confidence_scores[field] = confidence

    def _calculate_avg_confidence(self) -> float:
        """Calculate average confidence across all extracted fields"""
        if not self.confidence_scores:
            return 0.0
        return sum(self.confidence_scores.values()) / len(self.confidence_scores)

    def extract_billing_phases(self) -> List[Dict[str, Any]]:
        """
        Extract billing phases from scope of work section

        Returns:
            List of billing phase dicts
        """
        phases = []

        # Find BASIC SERVICES or SCOPE OF WORK section
        scope_start_idx = None
        scope_end_idx = None

        for i, line in enumerate(self.lines):
            line_upper = line.strip().upper()
            if 'BASIC SERVICES' in line_upper or 'SCOPE OF WORK' in line_upper:
                scope_start_idx = i
            elif scope_start_idx is not None and ('ADDITIONAL SERVICES' in line_upper or 'COMPENSATION' in line_upper):
                scope_end_idx = i
                break

        if scope_start_idx is None:
            return phases

        # Extract scope text
        if scope_end_idx is None:
            scope_end_idx = min(scope_start_idx + 50, len(self.lines))  # Max 50 lines

        scope_text = '\n'.join(self.lines[scope_start_idx:scope_end_idx])

        # Match phase keywords
        for phase_name, keywords in self.PHASE_KEYWORDS.items():
            for keyword in keywords:
                if keyword.lower() in scope_text.lower():
                    # Try to extract dollar amount associated with this phase
                    # Look for dollar amount within 100 chars of keyword
                    keyword_pos = scope_text.lower().find(keyword.lower())
                    context = scope_text[keyword_pos:keyword_pos+200]
                    amount_match = re.search(r'\$\s*([\d,]+(?:\.\d{2})?)', context)

                    amount = None
                    if amount_match:
                        amount_str = amount_match.group(1).replace(',', '')
                        amount = str(Decimal(amount_str))

                    phases.append({
                        'phase_name': phase_name,
                        'amount': amount,
                        'order': len(phases),
                    })
                    break  # Only add phase once

        return phases


class PIFExcelGenerator:
    """
    Generates PIF Excel files matching the format expected by pif_parser.py
    """

    # Layout modes (matching parser expectations)
    MODE_DEFAULT = 1        # Col 0 = label, Col 1 = value
    MODE_CLIENT_SECTION = 2 # Col 0 = header, Col 1 = label, Col 2 = value
    MODE_PROJECT_SECTION = 3 # Col 0 = label, Col 1 = value (same as DEFAULT)
    MODE_TABLE = 4          # Table structure for billing phases

    def __init__(self):
        """Initialize the Excel generator"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")

        self.workbook = None
        self.sheet = None
        self.current_row = 1

    def generate(self, pif_data: Dict[str, Any], billing_phases: List[Dict[str, Any]] = None) -> Workbook:
        """
        Generate PIF Excel workbook

        Args:
            pif_data: Dictionary of PIF field values
            billing_phases: List of billing phase dicts

        Returns:
            openpyxl Workbook
        """
        self.workbook = Workbook()

        # Remove default sheet and create Cover sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

        # Create Cover sheet (optional, simple title page)
        cover_sheet = self.workbook.create_sheet('Cover', 0)
        cover_sheet['A1'] = 'PROJECT INFORMATION FORM'
        cover_sheet['A1'].font = Font(size=18, bold=True)

        # Create PIF Data sheet (main data)
        self.sheet = self.workbook.create_sheet('PIF Data', 1)
        self.current_row = 1

        # Build the sheet
        self._add_title()
        self._add_project_identification_section(pif_data)
        self._add_client_information_section(pif_data)
        self._add_project_management_section(pif_data)
        self._add_special_tax_section(pif_data)
        self._add_special_billing_section(pif_data)
        self._add_billing_phases_section(billing_phases or [])

        # Set column widths
        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 30
        self.sheet.column_dimensions['C'].width = 30

        return self.workbook

    def _add_title(self):
        """Add title row"""
        self.current_row = 2
        self.sheet[f'A{self.current_row}'] = 'PROJECT INFORMATION FORM'
        self.sheet[f'A{self.current_row}'].font = Font(size=14, bold=True)
        self.sheet[f'A{self.current_row}'].alignment = Alignment(horizontal='center')
        self.current_row += 2  # Skip a row

    def _add_section_header(self, title: str):
        """Add bold section header"""
        self.sheet[f'A{self.current_row}'] = title
        self.sheet[f'A{self.current_row}'].font = Font(bold=True, size=12)
        self.current_row += 1

    def _add_field(self, label: str, value: Any, mode: int = MODE_DEFAULT):
        """
        Add a field to the sheet based on layout mode

        Args:
            label: Field label
            value: Field value
            mode: Layout mode
        """
        if mode == self.MODE_CLIENT_SECTION:
            # MODE_CLIENT_SECTION: Column 0 = section header (already added), Column 1 = label, Column 2 = value
            # Note: Column 0 stays blank for field rows, it was set by _add_section_header
            self.sheet[f'B{self.current_row}'] = label
            self.sheet[f'C{self.current_row}'] = value if value else ''
        else:
            # MODE_DEFAULT or MODE_PROJECT_SECTION
            # Column 0 = label, Column 1 = value
            self.sheet[f'A{self.current_row}'] = label
            self.sheet[f'B{self.current_row}'] = value if value else ''

        self.current_row += 1

    def _add_project_identification_section(self, pif_data: Dict[str, Any]):
        """Add Project Identification section (MODE_DEFAULT)"""
        self._add_field('Project Number', pif_data.get('project_number'), self.MODE_DEFAULT)
        self._add_field('Project Name', pif_data.get('project_name'), self.MODE_DEFAULT)
        self._add_field('DLAA Office', pif_data.get('dlaa_office'), self.MODE_DEFAULT)
        self._add_field('City', pif_data.get('project_location_city'), self.MODE_DEFAULT)
        self._add_field('State', pif_data.get('project_location_state'), self.MODE_DEFAULT)
        self._add_field('Originator', pif_data.get('originator'), self.MODE_DEFAULT)
        self._add_field('Date Entered', pif_data.get('date_entered'), self.MODE_DEFAULT)
        self.current_row += 1  # Skip row before next section

    def _add_client_information_section(self, pif_data: Dict[str, Any]):
        """Add Client Information section (MODE_CLIENT_SECTION)"""
        self._add_section_header('CLIENT INFORMATION')
        # Use exact label from parser: "Name (Firm or Individual)"
        self._add_field('Name (Firm or Individual)', pif_data.get('client_name'), self.MODE_CLIENT_SECTION)
        self._add_field('Billing Contact', pif_data.get('billing_contact'), self.MODE_CLIENT_SECTION)
        self._add_field('Billing Contact Email', pif_data.get('billing_contact_email'), self.MODE_CLIENT_SECTION)
        self._add_field('Phone', pif_data.get('phone'), self.MODE_CLIENT_SECTION)
        # Add Client Code if available (parser expects this)
        if pif_data.get('client_code'):
            self._add_field('Client Code', pif_data.get('client_code'), self.MODE_CLIENT_SECTION)
        # Add address fields if available
        if pif_data.get('billing_address_line_1'):
            self._add_field('Address Line 1', pif_data.get('billing_address_line_1'), self.MODE_CLIENT_SECTION)
        if pif_data.get('billing_address_line_2'):
            self._add_field('Address Line 2', pif_data.get('billing_address_line_2'), self.MODE_CLIENT_SECTION)
        if pif_data.get('billing_zip'):
            self._add_field('Zip', pif_data.get('billing_zip'), self.MODE_CLIENT_SECTION)
        self._add_field('Secondary Contact', pif_data.get('secondary_contact'), self.MODE_CLIENT_SECTION)
        self._add_field('Secondary Contact Email', pif_data.get('secondary_contact_email'), self.MODE_CLIENT_SECTION)
        self._add_field('Client Project Name', pif_data.get('client_project_name'), self.MODE_CLIENT_SECTION)
        self._add_field('Purchase Order Number', pif_data.get('purchase_order_number'), self.MODE_CLIENT_SECTION)
        self.current_row += 1

    def _add_project_management_section(self, pif_data: Dict[str, Any]):
        """Add Project Management section (MODE_PROJECT_SECTION - triggers mode change in parser)"""
        # Don't add a section header here - "Project Manager" in column 0 triggers the mode change
        # This is MODE_PROJECT_SECTION which is same layout as MODE_DEFAULT (col 0 = label, col 1 = value)
        # But the parser detects the transition when it sees "Project Manager" in column 0
        self._add_field('Project Manager', pif_data.get('project_manager'), self.MODE_DEFAULT)
        self._add_field('Project Start Date', pif_data.get('project_start_date'), self.MODE_DEFAULT)
        self._add_field('Fee (Contract Amount)', pif_data.get('fee_contract_amount'), self.MODE_DEFAULT)
        self._add_field('Type of Contract', pif_data.get('type_of_contract'), self.MODE_DEFAULT)
        self._add_field('Expenses', pif_data.get('expenses'), self.MODE_DEFAULT)
        self.current_row += 1

    def _add_special_tax_section(self, pif_data: Dict[str, Any]):
        """Add Special Tax Requirements section"""
        self._add_section_header('SPECIAL TAX REQUIREMENTS')

        # Tax locations checkboxes
        tax_locations = pif_data.get('tax_locations', [])
        if isinstance(tax_locations, str):
            # If stored as string, try to parse
            import json
            try:
                tax_locations = json.loads(tax_locations)
            except:
                tax_locations = []

        tax_options = ['Hawaii', 'Oahu', 'Maui', 'Kauai', 'Japan', 'South Korea', 'Guam', 'Singapore']
        tax_row = ' | '.join([f"{opt}: {'X' if opt.lower() in [t.lower() for t in tax_locations] else ''}" for opt in tax_options])
        self.sheet[f'A{self.current_row}'] = tax_row
        self.current_row += 2

    def _add_special_billing_section(self, pif_data: Dict[str, Any]):
        """Add Special Billing Information section"""
        self._add_section_header('SPECIAL BILLING INFORMATION')
        self._add_field('Special Negotiated Rates', pif_data.get('special_negotiated_rates'), self.MODE_DEFAULT)
        self._add_field('Special Invoice Instructions', pif_data.get('special_invoice_instructions'), self.MODE_DEFAULT)
        self._add_field('Retainer Received', pif_data.get('retainer_received'), self.MODE_DEFAULT)
        self._add_field('Additional Comments', pif_data.get('additional_comments'), self.MODE_DEFAULT)
        self.current_row += 1

    def _add_billing_phases_section(self, billing_phases: List[Dict[str, Any]]):
        """Add Billing Phases table (MODE_TABLE)"""
        self._add_section_header('BILLING PHASES')

        # Table headers
        headers = ['Phase', 'Max Amount', 'Amount', 'Subconsultant 1', 'Subconsultant 2']
        for col_idx, header in enumerate(headers):
            cell = self.sheet.cell(row=self.current_row, column=col_idx+1, value=header)
            cell.font = Font(bold=True)
        self.current_row += 1

        # Phase rows
        for phase in billing_phases:
            phase_name = phase.get('phase_name', '').replace('_', ' ').title()
            self.sheet.cell(row=self.current_row, column=1, value=phase_name)
            self.sheet.cell(row=self.current_row, column=2, value=phase.get('max_amount', ''))
            self.sheet.cell(row=self.current_row, column=3, value=phase.get('amount', ''))
            self.sheet.cell(row=self.current_row, column=4, value=phase.get('subconsultant_fee_1', ''))
            self.sheet.cell(row=self.current_row, column=5, value=phase.get('subconsultant_fee_2', ''))
            self.current_row += 1

    def save(self, filepath: str):
        """Save workbook to file"""
        self.workbook.save(filepath)
        logger.info(f"PIF Excel file saved to {filepath}")


class PIFGenerator:
    """
    Orchestrator class that coordinates extraction and Excel generation
    """

    def __init__(self, proposal_file_path: str):
        """
        Initialize the PIF generator

        Args:
            proposal_file_path: Path to proposal file (.doc/.docx/.txt)
        """
        self.proposal_path = proposal_file_path
        self.extracted_text = ""
        self.extracted_data = {}
        self.confidence_scores = {}
        self.billing_phases = []

    def extract_from_proposal(self) -> Tuple[Dict[str, Any], Dict[str, float]]:
        """
        Step 1: Extract text and parse fields from proposal

        Returns:
            Tuple of (extracted_data dict, confidence_scores dict)
        """
        logger.info(f"Extracting data from proposal: {self.proposal_path}")

        # Extract text using existing utility
        self.extracted_text = extract_text_from_file(self.proposal_path)

        if not self.extracted_text:
            logger.error(f"Failed to extract text from {self.proposal_path}")
            return {}, {}

        logger.info(f"Extracted {len(self.extracted_text)} characters of text")

        # Create field extractor and extract all fields
        extractor = PIFFieldExtractor(self.extracted_text, self.proposal_path)
        self.extracted_data, self.confidence_scores = extractor.extract_all_fields()

        # Extract billing phases
        self.billing_phases = extractor.extract_billing_phases()
        logger.info(f"Extracted {len(self.billing_phases)} billing phases")

        return self.extracted_data, self.confidence_scores

    def generate_excel(self, output_path: str, overrides: Optional[Dict[str, Any]] = None) -> str:
        """
        Step 2: Generate PIF Excel file

        Args:
            output_path: Path where Excel file should be saved
            overrides: Optional dict of field values to override extracted data

        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating PIF Excel file: {output_path}")

        # Merge extracted data with overrides
        pif_data = self.extracted_data.copy()
        if overrides:
            pif_data.update({k: v for k, v in overrides.items() if v})  # Only override non-empty values

        # Create Excel generator
        generator = PIFExcelGenerator()
        workbook = generator.generate(pif_data, self.billing_phases)

        # Save to file
        generator.save(output_path)

        return output_path

    def validate_extraction(self, excel_path: str) -> Dict[str, Any]:
        """
        Step 3: Validate generated Excel by parsing it with existing parser

        Args:
            excel_path: Path to generated Excel file

        Returns:
            Validation report dict
        """
        logger.info(f"Validating generated PIF: {excel_path}")

        try:
            # Import parser (avoid circular import)
            from billing.utils.pif_parser import parse_pif_excel

            # Parse the generated Excel
            parsed_data = parse_pif_excel(excel_path, use_second_sheet_only=True)

            # Compare extracted fields to parsed fields
            mismatches = []
            for field, original_value in self.extracted_data.items():
                parsed_value = parsed_data.get(field)
                if str(original_value) != str(parsed_value):
                    mismatches.append({
                        'field': field,
                        'original': original_value,
                        'parsed': parsed_value,
                    })

            if mismatches:
                logger.warning(f"Found {len(mismatches)} field mismatches during validation")
                return {
                    'success': False,
                    'message': f"{len(mismatches)} field(s) did not round-trip correctly",
                    'mismatches': mismatches,
                }
            else:
                logger.info("Validation successful - all fields round-tripped correctly")
                return {
                    'success': True,
                    'message': "All fields validated successfully",
                }

        except Exception as e:
            logger.error(f"Validation error: {e}", exc_info=True)
            return {
                'success': False,
                'message': f"Validation error: {str(e)}",
            }
